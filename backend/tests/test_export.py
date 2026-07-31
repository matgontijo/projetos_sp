"""Exportacao: PDF, Excel e CSV do fechamento.

Sao a cara do trabalho para quem recebe o relatorio — e nada aqui e coberto pelos
testes de calculo, entao um erro de layout so aparecia na mao do usuario.
"""

import io

import pytest
from openpyxl import load_workbook

from app.services import export


def linha(nome: str, receita: float, resultado_negativo: bool = False, conf: dict | None = None) -> dict:
    producao = receita * (1.4 if resultado_negativo else 0.4)
    custo = producao + 1000.0
    resultado = receita - custo
    return {
        "projeto": nome,
        "empresas": "CHERRY, JPDV",
        "cliente": "CLIENTE DE TESTE COM RAZÃO SOCIAL BEM LONGA LTDA ME",
        "receita": receita,
        "producao": producao,
        "frete": 500.0,
        "comissao": 300.0,
        "outros": 200.0,
        "imposto": 1000.0,
        "custo_total": custo,
        "resultado": resultado,
        "margem": resultado / receita if receita else 0,
        "conferencia": conf
        or {
            "status": "pendente",
            "oks": 0,
            "conferido_por": "",
            "conferido_em": None,
            "aprovado_por": "",
            "aprovado_em": None,
            "divergente": False,
        },
    }


APROVADO = {
    "status": "aprovado",
    "oks": 2,
    "conferido_por": "Maria",
    "conferido_em": "2026-07-30T18:12:00-03:00",
    "aprovado_por": "João",
    "aprovado_em": "2026-07-30T19:40:00-03:00",
    "divergente": False,
}


@pytest.fixture()
def dados():
    projetos = [
        linha("BR26_001", 100_000.0, conf=APROVADO),
        linha("BR26_002", 50_000.0),
        linha("BR26_003", 10_000.0, resultado_negativo=True),
    ]
    receita = sum(p["receita"] for p in projetos)
    resultado = sum(p["resultado"] for p in projetos)
    consolidado = {
        "receita": receita,
        "producao": sum(p["producao"] for p in projetos),
        "frete": sum(p["frete"] for p in projetos),
        "comissao": sum(p["comissao"] for p in projetos),
        "outros": sum(p["outros"] for p in projetos),
        "imposto": sum(p["imposto"] for p in projetos),
        "custo_total": sum(p["custo_total"] for p in projetos),
        "resultado": resultado,
        "margem_media": resultado / receita,
        "qtd_projetos": len(projetos),
        "cp_impostos": 2_500.0,
        "nao_classificado": 900.0,
        "qtd_aprovados": 1,
        "qtd_conferidos": 0,
        "qtd_pendentes": 2,
        "qtd_divergentes": 0,
    }
    return projetos, consolidado


# --- PDF ---


def test_pdf_sai_valido_e_com_fonte_embutida(dados):
    conteudo = export.fechamento_pdf(*dados, "Período: 01/01/2026 a 30/07/2026")

    assert conteudo.startswith(b"%PDF")
    assert conteudo.rstrip().endswith(b"%%EOF")
    # Manrope embutida: sem ela o PDF cai em fonte core, que come acento
    assert b"Manrope" in conteudo
    assert len(conteudo) > 10_000


def test_pdf_pagina_muitos_projetos(dados):
    _, consolidado = dados
    muitos = [linha(f"BR26_{i:03d}", 10_000.0 + i) for i in range(120)]
    conteudo = export.fechamento_pdf(muitos, consolidado, "")

    assert conteudo.count(b"/Type /Page\n") >= 3  # quebrou em varias paginas


def test_pdf_aguenta_lista_vazia(dados):
    _, consolidado = dados
    assert export.fechamento_pdf([], {**consolidado, "qtd_projetos": 0}, "").startswith(b"%PDF")


def test_marca_de_conferencia_no_pdf():
    assert export._conf_curto({"conferencia": {"oks": 0}}) == "—"
    assert export._conf_curto({"conferencia": {"oks": 1}}) == "1/2"
    assert export._conf_curto({"conferencia": {"oks": 2}}) == "2/2"
    assert export._conf_curto({"conferencia": {"oks": 2, "divergente": True}}) == "2/2!"
    assert export._conf_curto({}) == "—"  # linha sem bloco de conferência


# --- Excel ---


def test_xlsx_e_planilha_de_trabalho(dados):
    projetos, consolidado = dados
    wb = load_workbook(io.BytesIO(export.fechamento_xlsx(projetos, consolidado)))

    assert wb.sheetnames == ["Fechamento", "Resumo"]
    ws = wb["Fechamento"]
    # cabecalho fixo e filtro: a pessoa rola e filtra sem formatar nada
    assert ws.freeze_panes == "D2"
    assert ws.auto_filter.ref == f"A1:R{len(projetos) + 1}"
    assert ws.max_row == len(projetos) + 2  # cabecalho + projetos + total


def test_xlsx_formata_moeda_e_percentual(dados):
    projetos, consolidado = dados
    ws = load_workbook(io.BytesIO(export.fechamento_xlsx(projetos, consolidado)))["Fechamento"]

    assert "[Red]" in ws.cell(row=2, column=4).number_format  # negativo em vermelho
    assert ws.cell(row=2, column=12).number_format == "0.0%"
    # numero de verdade, nao texto: e o que deixa somar e ordenar no Excel
    # (o openpyxl relê 100000.0 como int quando nao ha casa decimal)
    assert isinstance(ws.cell(row=2, column=4).value, (int, float))
    assert isinstance(ws.cell(row=2, column=12).value, float)


def test_xlsx_leva_a_conferencia_e_o_total(dados):
    projetos, consolidado = dados
    ws = load_workbook(io.BytesIO(export.fechamento_xlsx(projetos, consolidado)))["Fechamento"]

    assert ws.cell(row=1, column=13).value == "Conferência"
    assert ws.cell(row=2, column=13).value == "Conferido e aprovado"
    assert ws.cell(row=2, column=14).value == "Maria"
    assert ws.cell(row=3, column=13).value == "Pendente"

    total = ws[ws.max_row]
    assert total[0].value == "TOTAL"
    assert total[3].value == pytest.approx(consolidado["receita"])
    assert "1 de 3" in total[12].value  # quantos fecharam a conferência


def test_aba_resumo_responde_o_periodo(dados):
    projetos, consolidado = dados
    ws = load_workbook(io.BytesIO(export.fechamento_xlsx(projetos, consolidado)))["Resumo"]
    valores = {row[0]: row[1] for row in ws.iter_rows(values_only=True) if row[0]}

    assert valores["Receita"] == pytest.approx(consolidado["receita"])
    assert valores["Resultado"] == pytest.approx(consolidado["resultado"])
    assert valores["Conferidos e aprovados (2 ok)"] == 1
    assert valores["Tributos pagos via contas a pagar"] == pytest.approx(2_500.0)


# --- CSV ---


def test_csv_pt_br_com_conferencia(dados):
    projetos, consolidado = dados
    texto = export.fechamento_csv(projetos, consolidado)
    linhas = texto.splitlines()

    assert linhas[0].startswith("Empresas;Projeto;Cliente")
    assert "Conferência" in linhas[0]
    assert "1º ok (conferiu)" in linhas[0]
    # decimal com virgula, separador ';' — o Excel pt-BR abre direto
    assert "100000,00" in linhas[1]
    assert "Conferido e aprovado;Maria;30/07/2026 18:12" in linhas[1]
    assert linhas[-1].startswith("TOTAL")


def test_csv_nao_quebra_sem_bloco_de_conferencia(dados):
    projetos, consolidado = dados
    for p in projetos:
        p.pop("conferencia")

    linhas = export.fechamento_csv(projetos, consolidado).splitlines()
    assert "Pendente" in linhas[1]
