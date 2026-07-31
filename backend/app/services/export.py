"""Exportacao do fechamento em CSV (pt-BR, ';'), Excel (openpyxl) e PDF (fpdf2).

O PDF segue a mesma linguagem editorial da proposta comercial: Manrope embutida
(nada de fonte core, que so aguenta Latin-1 e come acento), faixa preta com o
logotipo do grupo no topo, fios finos no lugar de grades pesadas e o resultado
como heroi da pagina.
"""

import csv
import io
from datetime import datetime, timedelta, timezone
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# a mesma familia do app; ficam no pacote de precificacao por historico
_FONTES = Path(__file__).resolve().parent.parent / "precificacao" / "fonts"
TZ_BR = timezone(timedelta(hours=-3))  # o servidor roda em UTC


def _agora_br() -> datetime:
    return datetime.now(TZ_BR)


# paleta da marca (preto e branco), a mesma da proposta e do app
from ..marca import (  # noqa: E402  (paleta antes das constantes que a usam)
    CINZA,
    CINZA_XL,
    CLARO,
    FIO,
    FIO_XL,
    NOME,
    PRETO,
    PRETO_XL,
    TINTA,
    VERDE,
    VERMELHO,
    ZEBRA,
    ZEBRA_XL,
    desenhar_logotipo,
)

COLUNAS = [
    ("empresas", "Empresas"),
    ("projeto", "Projeto"),
    ("cliente", "Cliente"),
    ("receita", "Receita (R$)"),
    ("producao", "Produção (R$)"),
    ("frete", "Frete (R$)"),
    ("comissao", "Comissão (R$)"),
    ("imposto", "Impostos (R$)"),
    ("outros", "Outros (R$)"),
    ("custo_total", "Custo total (R$)"),
    ("resultado", "Resultado (R$)"),
    ("margem", "Margem (%)"),
]

# Dupla conferencia — colunas de texto, sempre DEPOIS das numericas (o formato
# de moeda/percentual do Excel e aplicado por posicao nas 12 primeiras).
COLUNAS_CONFERENCIA = [
    "Conferência",
    "1º ok (conferiu)",
    "Conferido em",
    "2º ok (aprovou)",
    "Aprovado em",
    "Mudou após conferência",
]

_STATUS_LEGIVEL = {
    "pendente": "Pendente",
    "conferido": "Conferido (falta aprovar)",
    "aprovado": "Conferido e aprovado",
}


def _data_hora_pt(iso: str | None) -> str:
    if not iso:
        return ""
    try:
        return datetime.fromisoformat(iso).strftime("%d/%m/%Y %H:%M")
    except ValueError:
        return str(iso)


def _valores_conferencia(linha: dict) -> list[str]:
    conf = linha.get("conferencia") or {}
    return [
        _STATUS_LEGIVEL.get(conf.get("status", ""), "Pendente"),
        conf.get("conferido_por") or "",
        _data_hora_pt(conf.get("conferido_em")),
        conf.get("aprovado_por") or "",
        _data_hora_pt(conf.get("aprovado_em")),
        "Sim" if conf.get("divergente") else "",
    ]


def _conf_curto(linha: dict) -> str:
    """Marca compacta do PDF: '—', '1/2', '2/2' — com '!' se mudou depois do ok."""
    conf = linha.get("conferencia") or {}
    oks = int(conf.get("oks") or 0)
    marca = f"{oks}/2" if oks else "—"
    return f"{marca}!" if conf.get("divergente") else marca


def _valor_pt_br(campo: str, valor) -> str:
    if campo == "margem":
        return f"{valor * 100:.2f}".replace(".", ",")
    if isinstance(valor, (int, float)) and campo not in ("empresa_id", "codigo_projeto"):
        return f"{valor:.2f}".replace(".", ",")
    return str(valor)


def fechamento_csv(projetos: list[dict], consolidado: dict) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\r\n")
    writer.writerow([titulo for _, titulo in COLUNAS] + COLUNAS_CONFERENCIA)
    for linha in projetos:
        writer.writerow(
            [_valor_pt_br(campo, linha.get(campo, "")) for campo, _ in COLUNAS] + _valores_conferencia(linha)
        )
    writer.writerow([])
    writer.writerow(
        ["TOTAL", "", ""]
        + [
            _valor_pt_br(campo, consolidado.get("margem_media" if campo == "margem" else campo, 0))
            for campo, _ in COLUNAS[3:]
        ]
        + [f"{consolidado.get('qtd_pendentes', 0)} pendente(s)"]
        + [""] * (len(COLUNAS_CONFERENCIA) - 1)
    )
    return buffer.getvalue()


def _pdf_txt(valor) -> str:
    """Mantido por compatibilidade — com Manrope embutida o PDF e UTF-8."""
    return str(valor)


def _moeda_pt(valor: float) -> str:
    inteiro, decimal = f"{abs(valor):,.2f}".split(".")
    inteiro = inteiro.replace(",", ".")
    sinal = "-" if valor < 0 else ""
    return f"{sinal}{inteiro},{decimal}"


def _moeda_curta(valor: float) -> str:
    """Valor grande em forma legivel para o painel de KPI (2,1 mi / 340 mil)."""
    absoluto = abs(valor)
    sinal = "-" if valor < 0 else ""
    if absoluto >= 1_000_000:
        return f"{sinal}{absoluto / 1_000_000:,.1f} mi".replace(".", ",")
    if absoluto >= 1_000:
        return f"{sinal}{absoluto / 1_000:,.0f} mil".replace(",", ".")
    return f"{sinal}{absoluto:,.0f}".replace(",", ".")


# (campo, titulo, largura mm, alinhamento) — soma = 273mm, a util do A4 paisagem
_COLUNAS_PDF = [
    ("projeto", "Projeto", 40, "L"),
    ("_conf", "Conf.", 11, "C"),
    ("empresas", "Empresas", 20, "L"),
    ("cliente", "Cliente", 33, "L"),
    ("receita", "Receita", 26, "R"),
    ("producao", "Produção", 23, "R"),
    ("frete", "Frete", 18, "R"),
    ("imposto", "Impostos", 23, "R"),
    ("comissao", "Comissão", 20, "R"),
    ("outros", "Outros", 18, "R"),
    ("resultado", "Resultado", 26, "R"),
    ("margem", "Margem", 15, "R"),
]


class _RelatorioBase(FPDF):
    """Faixa de marca no topo e rodape paginado em todas as paginas."""

    def __init__(self, subtitulo: str):
        super().__init__(orientation="L", format="A4")
        self.subtitulo = subtitulo
        self.set_margins(12, 10, 12)
        self.set_auto_page_break(auto=True, margin=16)
        self.c_margin = 1.2
        self.add_font("Manrope", "", _FONTES / "Manrope-Regular.ttf")
        self.add_font("Manrope", "B", _FONTES / "Manrope-Bold.ttf")
        self.add_font("ManropeX", "", _FONTES / "Manrope-ExtraBold.ttf")

    def header(self):
        util = self.w - self.l_margin - self.r_margin
        self.set_fill_color(*PRETO)
        self.rect(0, 0, self.w, 26, style="F")

        # logotipo do grupo a esquerda
        largura_marca = desenhar_logotipo(self, self.l_margin, 6.5, tamanho=15)

        # titulo do documento, logo depois da marca (separado por um fio vertical)
        x_titulo = self.l_margin + largura_marca + 8
        self.set_draw_color(*CINZA)
        self.set_line_width(0.2)
        self.line(x_titulo - 4, 7, x_titulo - 4, 19)
        self.set_xy(x_titulo, 8)
        self.set_font("ManropeX", "", 12)
        self.set_text_color(255, 255, 255)
        self.cell(util - largura_marca - 98, 6, "Fechamento de Projetos")
        self.set_xy(x_titulo, 14.5)
        self.set_font("Manrope", "", 6.8)
        self.set_text_color(*CLARO)
        self.cell(util - largura_marca - 98, 4, "Receita − Produção − Frete − Comissão − Impostos − Outros = Resultado")

        if self.subtitulo:
            self.set_xy(self.w - self.r_margin - 90, 10)
            self.set_font("Manrope", "B", 8)
            self.set_text_color(*CLARO)
            self.cell(90, 6, self.subtitulo, align="R")
        self.set_y(33)

    def footer(self):
        util = self.w - self.l_margin - self.r_margin
        self.set_y(-12)
        self.set_draw_color(*FIO)
        self.set_line_width(0.2)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.set_y(-9.5)
        self.set_font("Manrope", "", 6.5)
        self.set_text_color(*CINZA)
        self.cell(
            util - 40,
            4,
            f"Grupo JPDV · somente projetos de venda (numeração BR) · "
            f"gerado em {_agora_br().strftime('%d/%m/%Y %H:%M')}",
        )
        self.cell(40, 4, f"Página {self.page_no()} de {{nb}}", align="R")


def _encurtar(pdf: FPDF, texto: str, largura: float) -> str:
    """Corta pela largura REAL do texto na fonte, nao por contagem de caracteres."""
    limite = largura - 2 * pdf.c_margin
    if pdf.get_string_width(texto) <= limite:
        return texto
    cortado = texto
    while cortado and pdf.get_string_width(cortado + "…") > limite:
        cortado = cortado[:-1]
    return (cortado + "…") if cortado else ""


def _kpi(pdf: FPDF, x: float, y: float, largura: float, rotulo: str, valor: str, destaque: bool = False) -> None:
    """Métrica editorial: tick preto, rótulo em caixa alta, valor grande."""
    pdf.set_fill_color(*PRETO)
    pdf.rect(x, y, 8, 1.3, style="F")
    pdf.set_xy(x, y + 3)
    pdf.set_font("ManropeX", "", 6.4)
    pdf.set_text_color(*CINZA)
    pdf.set_char_spacing(0.6)
    pdf.cell(largura, 4, rotulo.upper())
    pdf.set_char_spacing(0)
    pdf.set_xy(x, y + 7.5)
    pdf.set_font("ManropeX", "", 14 if destaque else 12)
    pdf.set_text_color(*TINTA)
    pdf.cell(largura, 7, valor)


def _cabecalho_tabela(pdf: FPDF) -> None:
    pdf.set_font("ManropeX", "", 6.6)
    pdf.set_text_color(*CINZA)
    pdf.set_draw_color(*PRETO)
    pdf.set_line_width(0.35)
    pdf.set_char_spacing(0.4)
    for _, titulo, largura, alinh in _COLUNAS_PDF:
        pdf.cell(largura, 6, titulo.upper(), border="B", align=alinh)
    pdf.set_char_spacing(0)
    pdf.ln()
    pdf.set_draw_color(*FIO)
    pdf.set_line_width(0.15)


def fechamento_pdf(projetos: list[dict], consolidado: dict, subtitulo: str = "") -> bytes:
    """Relatorio de fechamento (A4 paisagem): painel de KPI e uma linha por projeto."""
    pdf = _RelatorioBase(subtitulo)
    pdf.alias_nb_pages()
    pdf.add_page()
    util = pdf.w - pdf.l_margin - pdf.r_margin

    # ---- painel de KPI ----
    receita = float(consolidado.get("receita", 0) or 0)
    resultado = float(consolidado.get("resultado", 0) or 0)
    custos = sum(
        float(consolidado.get(c, 0) or 0) for c in ("producao", "frete", "comissao", "outros")
    )
    kpis = [
        ("Receita", f"R$ {_moeda_curta(receita)}", False),
        ("Custos", f"R$ {_moeda_curta(custos)}", False),
        ("Impostos", f"R$ {_moeda_curta(float(consolidado.get('imposto', 0) or 0))}", False),
        ("Resultado", f"R$ {_moeda_curta(resultado)}", True),
        ("Margem", f"{consolidado.get('margem_media', 0) * 100:.1f}%".replace(".", ","), True),
    ]
    y0 = pdf.get_y()
    passo = util / len(kpis)
    for i, (rotulo, valor, destaque) in enumerate(kpis):
        _kpi(pdf, pdf.l_margin + i * passo, y0, passo - 4, rotulo, valor, destaque)
    pdf.set_y(y0 + 17)

    # ---- linha de contexto: quantos projetos e como esta a conferencia ----
    total_projetos = consolidado.get("qtd_projetos", len(projetos))
    aprovados = consolidado.get("qtd_aprovados", 0)
    partes = [f"{total_projetos} projetos de venda", f"{aprovados} com os dois ok da conferência"]
    if consolidado.get("qtd_conferidos"):
        partes.append(f"{consolidado['qtd_conferidos']} esperando o 2º ok")
    if consolidado.get("qtd_divergentes"):
        partes.append(f"{consolidado['qtd_divergentes']} mudaram depois do ok")
    pdf.set_font("Manrope", "", 7.5)
    pdf.set_text_color(*CINZA)
    pdf.cell(util, 4.5, " · ".join(partes), new_x="LMARGIN", new_y="NEXT")
    pdf.ln(3)

    # ---- tabela ----
    _cabecalho_tabela(pdf)
    limite_y = pdf.h - 20
    for i, linha in enumerate(projetos):
        if pdf.get_y() > limite_y:
            pdf.add_page()
            _cabecalho_tabela(pdf)
        preenche = i % 2 == 1
        if preenche:
            pdf.set_fill_color(*ZEBRA)
        negativo = float(linha.get("resultado", 0) or 0) < 0
        for campo, _, largura, alinh in _COLUNAS_PDF:
            if campo == "_conf":
                pdf.set_font("Manrope", "B", 6.6)
                pdf.set_text_color(*CINZA)
                pdf.cell(largura, 5.6, _conf_curto(linha), border="B", align=alinh, fill=preenche)
                continue
            valor = linha.get(campo, "")
            if campo == "margem":
                texto = f"{float(valor or 0) * 100:.1f}%".replace(".", ",")
                pdf.set_font("Manrope", "B", 7)
            elif isinstance(valor, (int, float)):
                texto = _moeda_pt(float(valor))
                pdf.set_font("Manrope", "B" if campo == "resultado" else "", 7)
            else:
                pdf.set_font("Manrope", "B" if campo == "projeto" else "", 7)
                texto = _encurtar(pdf, str(valor), largura)
            if campo in ("resultado", "margem"):
                pdf.set_text_color(*(VERMELHO if negativo else VERDE))
            else:
                pdf.set_text_color(*TINTA)
            pdf.cell(largura, 5.6, texto, border="B", align=alinh, fill=preenche)
        pdf.ln()

    # ---- total ----
    if pdf.get_y() > limite_y - 8:
        pdf.add_page()
        _cabecalho_tabela(pdf)
    pdf.set_draw_color(*PRETO)
    pdf.set_line_width(0.35)
    totais = {
        "projeto": f"TOTAL · {len(projetos)} projeto(s)",
        "_conf": f"{aprovados}/{len(projetos)}",
        "empresas": "",
        "cliente": "",
        "receita": consolidado.get("receita", 0),
        "producao": consolidado.get("producao", 0),
        "frete": consolidado.get("frete", 0),
        "comissao": consolidado.get("comissao", 0),
        "imposto": consolidado.get("imposto", 0),
        "outros": consolidado.get("outros", 0),
        "resultado": consolidado.get("resultado", 0),
        "margem": consolidado.get("margem_media", 0),
    }
    for campo, _, largura, alinh in _COLUNAS_PDF:
        valor = totais.get(campo, "")
        pdf.set_font("ManropeX" if campo in ("resultado", "margem") else "Manrope", "" if campo in ("resultado", "margem") else "B", 7.2)
        if campo == "margem":
            texto = f"{float(valor or 0) * 100:.1f}%".replace(".", ",")
        elif isinstance(valor, (int, float)):
            texto = _moeda_pt(float(valor))
        else:
            texto = _encurtar(pdf, str(valor), largura)
        if campo in ("resultado", "margem"):
            pdf.set_text_color(*(VERMELHO if resultado < 0 else VERDE))
        else:
            pdf.set_text_color(*TINTA)
        pdf.cell(largura, 7.5, texto, border="T", align=alinh)
    pdf.ln(9)

    pdf.set_font("Manrope", "", 6.5)
    pdf.set_text_color(*CINZA)
    pdf.multi_cell(
        util,
        3.6,
        # sem "÷": a Manrope nao tem o glifo e o fpdf2 desenha outro caractere no lugar
        "Valores em R$. Margem = resultado / receita. Conf. = dupla conferência: 1/2 conferido, "
        "2/2 conferido e aprovado, “!” indica que os números mudaram depois do ok. "
        "Tributos pagos via contas a pagar não somam no custo — o imposto do projeto vem da NF-e.",
    )

    return bytes(pdf.output())


# --- Excel ---------------------------------------------------------------

_MOEDA_XL = 'R$ #,##0.00;[Red]-R$ #,##0.00'  # negativo em vermelho, como no Excel BR
_PCT_XL = "0.0%"
_MARCA_XL = PRETO_XL  # cabeçalhos e faixas de total, no preto da marca
_ZEBRA_XL = ZEBRA_XL
_FIO_XL = Side(style="thin", color=FIO_XL)


def _estilizar_cabecalho(ws, colunas: int) -> None:
    fundo = PatternFill("solid", fgColor=_MARCA_XL)
    for coluna in range(1, colunas + 1):
        cell = ws.cell(row=1, column=coluna)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = fundo
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


def fechamento_xlsx(projetos: list[dict], consolidado: dict) -> bytes:
    """Planilha de trabalho: cabecalho fixo, filtro, moeda BR e realce de margem.

    A pessoa abre e ja consegue filtrar, ordenar e enxergar prejuizo sem formatar
    nada — e a aba Resumo responde "como foi o periodo" sem rolar a tabela.
    """
    wb = Workbook()
    wb.properties.creator = f"{NOME} · Fechamento de Projetos"
    wb.properties.title = "Fechamento de projetos"
    ws = wb.active
    ws.title = "Fechamento"

    titulos = [titulo for _, titulo in COLUNAS] + COLUNAS_CONFERENCIA
    ws.append(titulos)
    _estilizar_cabecalho(ws, len(titulos))

    for i, linha in enumerate(projetos):
        ws.append(
            [
                linha.get(campo, "") if campo in ("empresas", "projeto", "cliente") else float(linha.get(campo, 0) or 0)
                for campo, _ in COLUNAS
            ]
            + _valores_conferencia(linha)
        )
        if i % 2 == 1:  # zebra: leitura de linha longa sem perder a linha
            zebra = PatternFill("solid", fgColor=_ZEBRA_XL)
            for cell in ws[ws.max_row]:
                cell.fill = zebra

    primeira, ultima = 2, ws.max_row  # faixa só de dados, sem o total

    ws.append(
        ["TOTAL", f"{len(projetos)} projeto(s)", ""]
        + [float(consolidado.get("margem_media" if campo == "margem" else campo, 0) or 0) for campo, _ in COLUNAS[3:]]
        + [f"{consolidado.get('qtd_aprovados', 0)} de {len(projetos)} com os dois ok"]
        + [""] * (len(COLUNAS_CONFERENCIA) - 1)
    )
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True, size=10)
        cell.border = Border(top=Side(style="medium", color=_MARCA_XL))

    for row in ws.iter_rows(min_row=2):
        for cell in row[3:11]:
            cell.number_format = _MOEDA_XL
        row[11].number_format = _PCT_XL
        for cell in row:
            cell.border = Border(bottom=_FIO_XL, top=cell.border.top)

    if ultima >= primeira:
        # resultado negativo salta aos olhos; margem ganha escala de cor
        ws.conditional_formatting.add(
            f"K{primeira}:K{ultima}",
            CellIsRule(operator="lessThan", formula=["0"], font=Font(bold=True, color="B72A2A")),
        )
        ws.conditional_formatting.add(
            f"L{primeira}:L{ultima}",
            ColorScaleRule(
                start_type="num", start_value=0, start_color="F8C9C9",
                mid_type="num", mid_value=0.2, mid_color="FFF4DB",
                end_type="num", end_value=0.5, end_color="CDEBD6",
            ),
        )
        # status da conferência (coluna M)
        for texto, cor in (
            ("Conferido e aprovado", "CDEBD6"),
            ("Conferido (falta aprovar)", "FFF4DB"),
            ("Pendente", "F1F2F4"),
        ):
            ws.conditional_formatting.add(
                f"M{primeira}:M{ultima}",
                CellIsRule(operator="equal", formula=[f'"{texto}"'], fill=PatternFill("solid", fgColor=cor)),
            )

    ws.freeze_panes = "D2"  # empresas, projeto e o selo ficam à vista ao rolar
    ws.auto_filter.ref = f"A1:{get_column_letter(len(titulos))}{ultima}"

    larguras = [22, 26, 30, 16, 16, 13, 14, 15, 13, 16, 16, 11, 24, 20, 17, 20, 17, 14]
    for idx, largura in enumerate(larguras[: len(titulos)], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = largura

    _aba_resumo(wb, projetos, consolidado)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _aba_resumo(wb: Workbook, projetos: list[dict], consolidado: dict) -> None:
    """Uma tela com o período inteiro: composição, margem e conferência."""
    ws = wb.create_sheet("Resumo")
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 14

    # marca no alto da folha, como no papel timbrado
    ws.append([NOME.upper()])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color=PRETO_XL)
    ws.append(["Fechamento de projetos"])
    ws.cell(row=2, column=1).font = Font(size=9, color=CINZA_XL)
    ws.append([])

    def titulo(texto: str) -> None:
        ws.append([texto])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=_MARCA_XL)
        ws.cell(row=ws.max_row, column=2).fill = PatternFill("solid", fgColor=_MARCA_XL)
        ws.cell(row=ws.max_row, column=3).fill = PatternFill("solid", fgColor=_MARCA_XL)

    def linha(rotulo: str, valor, formato: str | None = None, participacao: float | None = None) -> None:
        ws.append([rotulo, valor, participacao])
        if formato:
            ws.cell(row=ws.max_row, column=2).number_format = formato
        if participacao is not None:
            ws.cell(row=ws.max_row, column=3).number_format = _PCT_XL

    receita = float(consolidado.get("receita", 0) or 0)
    fatia = lambda v: (v / receita) if receita else None  # noqa: E731

    titulo("Resultado do período")
    linha("Receita", receita, _MOEDA_XL)
    for campo, rotulo in (
        ("producao", "Produção"),
        ("frete", "Frete"),
        ("comissao", "Comissão"),
        ("imposto", "Impostos"),
        ("outros", "Outros"),
    ):
        valor = float(consolidado.get(campo, 0) or 0)
        linha(rotulo, valor, _MOEDA_XL, fatia(valor))
    linha("Custo total", float(consolidado.get("custo_total", 0) or 0), _MOEDA_XL, fatia(float(consolidado.get("custo_total", 0) or 0)))
    linha("Resultado", float(consolidado.get("resultado", 0) or 0), _MOEDA_XL, fatia(float(consolidado.get("resultado", 0) or 0)))
    linha("Margem média", float(consolidado.get("margem_media", 0) or 0), _PCT_XL)
    for coluna in ("A", "B"):
        ws[f"{coluna}{ws.max_row}"].font = Font(bold=True)
        ws[f"{coluna}{ws.max_row - 1}"].font = Font(bold=True)

    ws.append([])
    titulo("Dupla conferência")
    linha("Projetos de venda", consolidado.get("qtd_projetos", len(projetos)))
    linha("Conferidos e aprovados (2 ok)", consolidado.get("qtd_aprovados", 0))
    linha("Esperando o 2º ok", consolidado.get("qtd_conferidos", 0))
    linha("Sem nenhum ok", consolidado.get("qtd_pendentes", 0))
    linha("Mudaram depois do ok", consolidado.get("qtd_divergentes", 0))

    ws.append([])
    titulo("Fora do custo (informativo)")
    linha("Tributos pagos via contas a pagar", float(consolidado.get("cp_impostos", 0) or 0), _MOEDA_XL)
    linha("Custos sem categoria mapeada", float(consolidado.get("nao_classificado", 0) or 0), _MOEDA_XL)

    ws.append([])
    ws.append(["Somente projetos de venda (numeração BR). Impostos vêm das NF-e; tributos"])
    ws.append(["do contas a pagar aparecem acima apenas como informação, sem somar no custo."])
    for deslocamento in (0, 1):
        ws.cell(row=ws.max_row - deslocamento, column=1).font = Font(italic=True, size=9, color="707682")
