"""PDF da proposta comercial + export do historico de orcamentos.

Design editorial: Manrope embutida (a mesma fonte do app), fios finos e
espaco em branco no lugar de caixas cheias. O total e o heroi da pagina,
a validade cria urgencia e o "proximo passo" fecha a venda.
"""

import csv
import io
from datetime import datetime, timedelta, timezone
from pathlib import Path

from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import Font

from ..services.export import _moeda_pt

_FONTES = Path(__file__).resolve().parent / "fonts"
TZ_BR = timezone(timedelta(hours=-3))  # Brasil sem horario de verao; servidor roda em UTC


def _agora_br() -> datetime:
    return datetime.now(TZ_BR)


def _para_br(dt: datetime) -> datetime:
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ_BR)


def _moeda4(valor: float) -> str:
    """4 casas quando 2 nao fecham a conta do cliente (unitario x quantidade)."""
    inteiro, decimal = f"{abs(valor):,.4f}".split(".")
    return f"{'-' if valor < 0 else ''}{inteiro.replace(',', '.')},{decimal}"

# paleta da proposta (impressao amigavel)
MARINHO = (15, 27, 51)
AZUL = (37, 99, 235)
ARDOSIA = (152, 168, 198)  # texto sobre o marinho
CINZA = (112, 118, 130)
TINTA = (18, 21, 27)
FIO = (228, 231, 236)  # hairlines
AMBAR_FUNDO = (255, 244, 219)
AMBAR_TEXTO = (140, 89, 8)


def _txt(valor) -> str:
    return str(valor)


def _marca(empresa_nome: str) -> str:
    """Nome fantasia curto a partir da razao social (2 primeiras palavras)."""
    palavras = (empresa_nome or "").split()
    return " ".join(palavras[:2]).upper() or "PROPOSTA"


class _PropostaBase(FPDF):
    """Cabecalho de marca e rodape com paginacao em todas as paginas."""

    def __init__(self, empresa_nome: str, numero: str, data: str):
        super().__init__(orientation="P", format="A4")
        self.empresa_nome = empresa_nome
        self.numero = numero
        self.data = data
        self.set_margins(14, 10, 14)
        self.set_auto_page_break(auto=True, margin=24)
        self.c_margin = 0  # texto rente as reguas/caixas: um unico eixo vertical no grid
        # a MESMA familia do app, embutida no arquivo
        self.add_font("Manrope", "", _FONTES / "Manrope-Regular.ttf")
        self.add_font("Manrope", "B", _FONTES / "Manrope-Bold.ttf")
        self.add_font("ManropeX", "", _FONTES / "Manrope-ExtraBold.ttf")

    def micro(self, texto: str, largura: float = 0, cor=CINZA, alinh: str = "L", altura: float = 4.5) -> None:
        """Micro-rotulo em caixa alta com espacamento de letras (assinatura editorial)."""
        self.set_font("ManropeX", "", 6.8)
        self.set_text_color(*cor)
        self.set_char_spacing(0.65)
        self.cell(largura, altura, _txt(texto.upper()), align=alinh)
        self.set_char_spacing(0)

    def header(self):
        util = self.w - self.l_margin - self.r_margin
        self.set_fill_color(*MARINHO)
        self.rect(0, 0, self.w, 30, style="F")
        # marca a esquerda
        self.set_xy(self.l_margin, 8)
        self.set_font("ManropeX", "", 17)
        self.set_text_color(255, 255, 255)
        self.cell(util - 72, 9, _txt(_marca(self.empresa_nome)))
        # chip do documento a direita
        chip_txt = f"PROPOSTA Nº {self.numero}"
        self.set_font("ManropeX", "", 8.5)
        w_chip = self.get_string_width(chip_txt) + 10
        x_chip = self.w - self.r_margin - w_chip
        self.set_fill_color(*AZUL)
        self.rect(x_chip, 8.2, w_chip, 8.6, style="F", round_corners=True, corner_radius=4.3)
        self.set_xy(x_chip, 8.2)
        self.set_text_color(255, 255, 255)
        self.cell(w_chip, 8.6, _txt(chip_txt), align="C")
        # segunda linha: razao social | data (mesma baseline)
        self.set_xy(self.l_margin, 20.5)
        self.set_font("Manrope", "", 7)
        self.set_text_color(*ARDOSIA)
        self.cell(util - 72, 4, _txt(self.empresa_nome[:80]))
        self.set_font("Manrope", "B", 7.5)
        self.cell(72, 4, _txt(f"Emitida em {self.data}"), align="R")
        self.set_y(40)

    def footer(self):
        util = self.w - self.l_margin - self.r_margin
        self.set_y(-14)
        self.set_draw_color(*FIO)
        self.set_line_width(0.2)
        self.line(self.l_margin, self.get_y(), self.w - self.r_margin, self.get_y())
        self.set_y(-11)
        self.set_font("Manrope", "", 6.5)
        self.set_text_color(*CINZA)
        gerado = _agora_br().strftime("%d/%m/%Y %H:%M")
        self.cell(util - 40, 4, _txt(f"{_marca(self.empresa_nome)} — proposta gerada em {gerado}"))
        self.cell(40, 4, _txt(f"Página {self.page_no()}"), align="R")


def proposta_pdf(orc, itens: list, empresa_nome: str) -> bytes:
    """Proposta comercial (A4, editorial): o total e o heroi, validade cria
    urgencia e o proximo passo diz ao cliente exatamente como aprovar."""
    criado_dt = _para_br(orc.criado_em) if orc.criado_em else _agora_br()
    validade = criado_dt + timedelta(days=15)
    condicao = "À vista" if orc.condicao_pagamento_dias == 0 else f"{orc.condicao_pagamento_dias} dias"
    qtd_total = sum(i["quantidade"] for i in itens) or 1
    # a soma das linhas e a verdade do documento: o cliente confere com calculadora
    total_doc = sum(float(i["total"]) for i in itens) or float(orc.total)

    pdf = _PropostaBase(empresa_nome, orc.numero, criado_dt.strftime("%d/%m/%Y"))
    pdf.add_page()
    util = pdf.w - pdf.l_margin - pdf.r_margin

    # ---- preparado para ----
    pdf.micro("Preparado para")
    pdf.ln(6)
    pdf.set_font("ManropeX", "", 16)
    pdf.set_text_color(*TINTA)
    pdf.multi_cell(util, 8.5, _txt(orc.cliente or "Cliente"), align="L", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(6)

    # ---- metricas editoriais: sem caixa; tick azul, rotulos e valores em baselines unicas ----
    y0 = pdf.get_y()
    pdf.set_fill_color(*AZUL)
    pdf.rect(pdf.l_margin, y0, 10, 1.5, style="F")
    X_FATOS = pdf.l_margin + 90
    LARG_FATO = 30.67  # 3 colunas terminam rentes a margem direita
    fatos = [
        ("Quantidade", f"{qtd_total:,}".replace(",", ".") + " un"),
        ("Preço unitário", f"R$ {_moeda_pt(total_doc / qtd_total)}"),
        ("Pagamento", condicao),
    ]
    # linha de rotulos
    pdf.set_xy(pdf.l_margin, y0 + 4.5)
    pdf.micro("Investimento total", 76)
    for i, (rotulo, _) in enumerate(fatos):
        pdf.set_xy(X_FATOS + i * LARG_FATO, y0 + 4.5)
        pdf.micro(rotulo, LARG_FATO - 2)
    # linha de valores: os menores descem 1,3mm para dividir a BASELINE com o numero-heroi
    pdf.set_xy(pdf.l_margin, y0 + 10)
    pdf.set_font("ManropeX", "", 23)
    pdf.set_text_color(*AZUL)
    pdf.cell(76, 11, _txt(f"R$ {_moeda_pt(total_doc)}"))
    pdf.set_font("Manrope", "B", 10.5)
    pdf.set_text_color(*TINTA)
    for i, (_, valor) in enumerate(fatos):
        pdf.set_xy(X_FATOS + i * LARG_FATO, y0 + 11.3)
        pdf.cell(LARG_FATO - 2, 11, _txt(valor))
    pdf.set_draw_color(*FIO)
    pdf.set_line_width(0.2)
    pdf.line(pdf.l_margin, y0 + 25.5, pdf.w - pdf.r_margin, y0 + 25.5)
    pdf.set_y(y0 + 30.5)

    # ---- validade: pill ambar do tamanho do texto ----
    aviso = f"Válida até {validade.strftime('%d/%m/%Y')} — a resposta dentro do prazo garante estas condições."
    pdf.set_font("Manrope", "B", 8.3)
    w_pill = pdf.get_string_width(aviso) + 11
    y0 = pdf.get_y()
    pdf.set_fill_color(*AMBAR_FUNDO)
    pdf.rect(pdf.l_margin, y0, min(w_pill, util), 8.4, style="F", round_corners=True, corner_radius=4.2)
    pdf.set_xy(pdf.l_margin, y0)
    pdf.set_text_color(*AMBAR_TEXTO)
    pdf.cell(min(w_pill, util), 8.4, _txt(aviso), align="C")
    pdf.set_y(y0 + 8.4 + 8)

    # ---- itens: tabela de fios finos (sem blocos de cor) ----
    pdf.micro("Itens da proposta")
    pdf.ln(6.5)
    colunas = [("item", "#", 8, "C"), ("descricao", "Descrição", 94, "L"),
               ("quantidade", "Qtde", 20, "R"), ("preco", "Preço unitário", 32, "R"), ("total", "Total", 28, "R")]
    for _, titulo, largura, alinh in colunas:
        pdf.micro(titulo, largura, alinh=alinh)
    pdf.ln(5.5)
    pdf.set_draw_color(*MARINHO)
    pdf.set_line_width(0.45)
    pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    pdf.set_draw_color(*FIO)
    pdf.set_line_width(0.2)
    pdf.set_font("Manrope", "", 9.5)
    for i, item in enumerate(itens, start=1):
        pdf.set_text_color(*TINTA)
        # unitario x quantidade TEM que bater com o total impresso (o cliente confere):
        # se 2 casas nao fecham a conta, exibe as 4 casas reais do unitario
        unit, qtd_i, tot_i = float(item["preco_unitario"]), item["quantidade"], float(item["total"])
        fecha_com_2_casas = abs(round(unit, 2) * qtd_i - tot_i) <= 0.01
        valores = {
            "item": str(i),
            "descricao": item["descricao"][:56],
            "quantidade": f"{qtd_i:,}".replace(",", "."),
            "preco": f"R$ {_moeda_pt(unit) if fecha_com_2_casas else _moeda4(unit)}",
            "total": f"R$ {_moeda_pt(tot_i)}",
        }
        for campo, _, largura, alinh in colunas:
            pdf.cell(largura, 9, _txt(valores[campo]), align=alinh)
        pdf.ln(9)
        pdf.line(pdf.l_margin, pdf.get_y(), pdf.w - pdf.r_margin, pdf.get_y())
    # total: par rotulo/valor rente a coluna da direita (rotulo desce p/ dividir a baseline)
    pdf.ln(2)
    y_tot = pdf.get_y()
    pdf.set_xy(pdf.l_margin + 98, y_tot + 1.1)
    pdf.micro("Total", 24, alinh="R", altura=10)
    pdf.set_xy(pdf.l_margin + 124, y_tot)
    pdf.set_font("ManropeX", "", 13.5)
    pdf.set_text_color(*AZUL)
    pdf.cell(util - 124, 10, _txt(f"R$ {_moeda_pt(total_doc)}"), align="R")
    pdf.set_y(y_tot + 10 + 7)

    # ---- condicoes em grade 2x2 ----
    pdf.micro("Condições de fornecimento")
    pdf.ln(6.5)
    pares = [
        ("Pagamento", condicao),
        ("Frete", "A combinar"),
        ("Faturamento por", _marca(empresa_nome)),
        ("Validade da proposta", validade.strftime("%d/%m/%Y")),
    ]
    meia = util / 2
    ROTULO_LARG = 36
    for linha in range(0, len(pares), 2):
        y_linha = pdf.get_y()
        for coluna, (rotulo, valor) in enumerate(pares[linha:linha + 2]):
            pdf.set_xy(pdf.l_margin + coluna * meia, y_linha)
            pdf.set_font("Manrope", "", 8)
            pdf.set_text_color(*CINZA)
            pdf.cell(ROTULO_LARG, 7, _txt(rotulo))
            pdf.set_font("Manrope", "B", 9)
            pdf.set_text_color(*TINTA)
            pdf.cell(meia - ROTULO_LARG, 7, _txt(valor))
        pdf.set_y(y_linha + 7)

    # ---- proximo passo: flui logo apos as condicoes (ancorar no pe abria um
    # vazio de ~90mm nas propostas curtas — critica de direcao de arte) ----
    ALTURA_CTA = 21
    pdf.ln(9)
    if pdf.get_y() + ALTURA_CTA > pdf.h - 24:  # nao deixa o cartao quebrar entre paginas
        pdf.add_page()
    y0 = pdf.get_y()
    pdf.set_fill_color(*MARINHO)
    pdf.rect(pdf.l_margin, y0, util, ALTURA_CTA, style="F", round_corners=True, corner_radius=3)
    pdf.set_xy(pdf.l_margin + 7, y0 + 4.5)
    pdf.micro("Próximo passo", cor=(255, 255, 255))
    pdf.set_xy(pdf.l_margin + 7, y0 + 11)
    pdf.set_font("Manrope", "", 8.7)
    pdf.set_text_color(198, 210, 232)
    responsavel = orc.criado_por or "nosso time comercial"
    pdf.cell(util - 14, 5, _txt(f"Para aprovar, responda esta proposta ou fale com {responsavel}. A produção inicia logo após a confirmação."))

    return bytes(pdf.output())


_COLS = [("numero", "Número"), ("cliente", "Cliente"), ("empresa", "Empresa"), ("status", "Status"),
         ("quantidade", "Qtde"), ("preco_unitario", "Preço unit. (R$)"), ("total", "Total (R$)"),
         ("condicao", "Condição"), ("criado_por", "Criado por"), ("criado_em", "Criado em")]


def _linhas_export(orcamentos: list[dict]) -> list[list]:
    return [[o.get(campo, "") for campo, _ in _COLS] for o in orcamentos]


def orcamentos_csv(orcamentos: list[dict]) -> str:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";", lineterminator="\r\n")
    writer.writerow([t for _, t in _COLS])
    for linha in _linhas_export(orcamentos):
        writer.writerow([str(v).replace(".", ",") if isinstance(v, float) else v for v in linha])
    return buffer.getvalue()


def orcamentos_xlsx(orcamentos: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamentos"
    ws.append([t for _, t in _COLS])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for linha in _linhas_export(orcamentos):
        ws.append(linha)
    for row in ws.iter_rows(min_row=2):
        row[5].number_format = 'R$ #,##0.0000'
        row[6].number_format = 'R$ #,##0.00'
    for idx, largura in enumerate([14, 30, 26, 12, 10, 16, 16, 12, 20, 18], start=1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = largura
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
